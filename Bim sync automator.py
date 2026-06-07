r"""
BIM File Sync Automator  v8.5  —  Light Blue Theme
Prepared by Ahmed Khalaf — BIM Manager
─────────────────────────────────────────────────────────────────────────────
Full white / light blue color scheme.
Copies .nwc  .ifc  .xlsx files.
All features preserved: multi-job, multi-source, pick-list, schedule,
import/export Excel & CSV.

CHANGE (v8.5): Clear handling of unavailable drives / network shares.
               - A disconnected mapped drive (e.g. Z:\ unavailable) now logs a
                 specific reason and tells you to reconnect it or use the
                 \\server\share UNC path, instead of a vague "missing folder".
               - Unreachable UNC shares and missing folders are reported
                 distinctly.
               - A job whose sources are all unreachable now logs a clear
                 warning instead of looking like a successful "Copied 0" run.
               - Destination on a disconnected drive is reported before the app
                 tries (and fails) to create folders on it.

CHANGE (v8.5): Overwrite of existing destination files is now reliable on
               Windows long paths and UNC shares.
               - Destination root is now \\?\ long-path prefixed (was source-only).
               - Removed the broken `src.lstrip("\\?\\")` line (lstrip strips a
                 CHARACTER SET, not a prefix — it mangled UNC paths). shutil.copy2
                 handles \\?\-prefixed paths directly and overwrites by default.

CHANGE (v8.3): Files now always overwrite existing files at the destination.
               The previous "skip if up-to-date" guard has been removed.

FIX: Long Windows path support (\\?\\ prefix) + os.makedirs on destination.
FIX: Log file now written to a per-user writable location instead of the
     installation directory (avoids PermissionError under C:\\Program Files).
─────────────────────────────────────────────────────────────────────────────
"""

import csv
import os
import shutil
import schedule
import time
import threading
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Optional, Set
import customtkinter as ctk
from tkinter import filedialog, messagebox
from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    # Define the names in the failure path too, so references elsewhere
    # resolve. They are only ever used when OPENPYXL_OK is True.
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None


# ── Writable log location ───────────────────────────────────────────────────────
def _resolve_log_path() -> str:
    """
    Return a log-file path inside a per-user writable directory.
    On Windows this is %LOCALAPPDATA%\\BIMSyncBuild\\logs, which is writable
    without admin rights — unlike the installation folder under Program Files.
    """
    if os.name == "nt":
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    else:
        base = os.environ.get("XDG_DATA_HOME") or os.path.join(
            os.path.expanduser("~"), ".local", "share"
        )
    log_dir = os.path.join(base, "BIMSyncBuild", "logs")
    try:
        os.makedirs(log_dir, exist_ok=True)
    except Exception:
        # Last-resort fallback: temp dir (always writable)
        import tempfile
        log_dir = os.path.join(tempfile.gettempdir(), "BIMSyncBuild")
        os.makedirs(log_dir, exist_ok=True)
    return os.path.join(log_dir, "bim_sync_history.log")


LOG_PATH = _resolve_log_path()

# ── Logging ────────────────────────────────────────────────────────────────────
logger.add(
    LOG_PATH,
    rotation="500 KB",
    retention=5,
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}",
)

# ── Constants ──────────────────────────────────────────────────────────────────
DEFAULT_PICKS: List[str] = [
    "WHZ", "WHS", "WHA", "WHB", "WHP",
    "WHG", "WBR", "WHM", "WHR", "WPB",
]
EXTENSIONS       = (".nwc", ".ifc", ".xlsx")
ALL_DAYS         = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_MAP          = {
    "Mon": "monday", "Tue": "tuesday", "Wed": "wednesday",
    "Thu": "thursday", "Fri": "friday", "Sat": "saturday", "Sun": "sunday",
}
COLUMNS = [
    "job_name", "sources", "destination", "exec_time", "days",
    "pick_list", "enabled", "filter_date", "auto_today",
]

# ── Light Blue Theme Colors ────────────────────────────────────────────────────
COLORS = {
    "green":    "#27ae60", "green_h":  "#219150",
    "red":      "#c0392b", "red_h":    "#a93226",
    "blue":     "#2980b9", "blue_h":   "#2471a3",
    "orange":   "#e67e22", "purple":   "#8e44ad",
    "teal":     "#16a085", "teal_h":   "#138d75",
    "gold":     "#b7950b", "gold_h":   "#9a7d0a",
    "dim":      "#5d7a9a",
    "frame_bg": "#d6e8f7",
    "row_bg":   "#eaf4fc",
    "pick_bg":  "#cde0f0",
    "src_bg":   "#daeaf7",
    "io_bar":   "#c5dcf0",
    "border":   "#a8c8e8",
    "text":     "#0d2b45",
    "text_dim": "#4a7a9b",
}

JOB_COLORS = [
    "#1a6fa8", "#6c3483", "#0e7a6e", "#c25b00",
    "#b03030", "#7d6608", "#1a6e3c", "#9b2060",
]


# ── Helpers ────────────────────────────────────────────────────────────────────
def validate_time(t: str) -> bool:
    try:
        datetime.strptime(t.strip(), "%H:%M")
        return True
    except ValueError:
        return False


def _split(s: str) -> List[str]:
    return [p.strip() for p in s.split("|") if p.strip()]


def _long_path(path: str) -> str:
    """Prefix path with \\?\\ on Windows to support paths longer than 260 chars."""
    if os.name == 'nt' and not path.startswith("\\\\?\\"):
        return "\\\\?\\" + os.path.abspath(path)
    return path


def _diagnose_path(path: str):
    """
    Return (accessible: bool, reason: str) for a folder path.

    Gives a SPECIFIC reason so the log explains the real problem instead of a
    generic "missing folder":
      - a disconnected mapped/removable drive  (e.g. Z:\\ unavailable)
      - an unreachable UNC network share        (e.g. \\\\server\\share down / no VPN)
      - a folder that simply does not exist
    """
    if not path or not path.strip():
        return False, "empty path"

    real = _long_path(path)
    if os.path.isdir(real):
        return True, "ok"

    # Drive-letter path like "Z:\\..." — check whether the drive root is there.
    drive, _ = os.path.splitdrive(path)
    if drive and len(drive) == 2 and drive.endswith(":"):
        root = drive + os.sep
        if not os.path.exists(root):
            return False, (f"drive {drive} is not connected — reconnect it, or use "
                           f"the \\\\server\\share UNC path instead of the {drive} letter")
        # Drive exists but the sub-folder doesn't.
        return False, "folder not found on that drive"

    # UNC network share like \\\\server\\share\\...
    if path.startswith("\\\\") or path.startswith("//"):
        return False, "network share unreachable — check the connection / VPN / path"

    return False, "folder not found"


# ── Job Data Model ─────────────────────────────────────────────────────────────
@dataclass
class SyncJob:
    job_id:       int
    job_name:     str       = ""
    sources:      List[str] = field(default_factory=list)
    dest:         str       = ""
    exec_time:    str       = "09:00"
    days:         List[str] = field(default_factory=lambda: ["Mon"])
    pick_list:    List[str] = field(default_factory=lambda: list(DEFAULT_PICKS))
    enabled:      bool      = True
    filter_date:  str       = ""
    auto_today:   bool      = False
    copied:       int       = 0
    failed:       int       = 0
    last_run:     str       = "—"

    @property
    def color(self) -> str:
        return JOB_COLORS[self.job_id % len(JOB_COLORS)]

    @property
    def label(self) -> str:
        return self.job_name or f"Job {self.job_id + 1}"

    @property
    def filter_datetime(self):
        if not self.filter_date:
            return None
        try:
            return datetime.strptime(self.filter_date.strip(), "%Y-%m-%d")
        except ValueError:
            return None

    def to_row(self) -> dict:
        return {
            "job_name":     self.label,
            "sources":      " | ".join(self.sources),
            "destination":  self.dest,
            "exec_time":    self.exec_time,
            "days":         " | ".join(self.days),
            "pick_list":    " | ".join(self.pick_list),
            "enabled":      "Yes" if self.enabled else "No",
            "filter_date":  self.filter_date,
            "auto_today":   "Yes" if self.auto_today else "No",
        }

    @staticmethod
    def from_row(row: dict, job_id: int) -> "SyncJob":
        enabled    = str(row.get("enabled",    "Yes")).strip().lower() not in ("no", "false", "0")
        auto_today = str(row.get("auto_today", "No" )).strip().lower() in ("yes", "true", "1")
        return SyncJob(
            job_id       = job_id,
            job_name     = row.get("job_name", "").strip(),
            sources      = _split(row.get("sources", "")),
            dest         = row.get("destination", "").strip(),
            exec_time    = row.get("exec_time", "09:00").strip() or "09:00",
            days         = _split(row.get("days", "Mon")) or ["Mon"],
            pick_list    = _split(row.get("pick_list", "")) or list(DEFAULT_PICKS),
            enabled      = enabled,
            filter_date  = row.get("filter_date", "").strip(),
            auto_today   = auto_today,
        )


# ── Excel Export ───────────────────────────────────────────────────────────────
def export_excel(jobs: List[SyncJob], filepath: str):
    if not OPENPYXL_OK:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sync Jobs"

    hdr_fill  = PatternFill
    hdr_font  = Font
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="B0B8C8")
    hdr_brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Job Name", "Source Folders\n(pipe-separated)", "Destination",
        "Exec Time\n(HH:MM)", "Days\n(pipe-separated)",
        "Pick-list Codes\n(pipe-separated)", "Enabled\n(Yes/No)",
        "Filter Date\n(YYYY-MM-DD)", "Auto Today\n(Yes/No)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = hdr_brd
    ws.row_dimensions[1].height = 36

    row_fills = ["EAF0FB", "F5F7FA"] * 8
    data_font  = Font(name="Consolas", size=10, color="1A1A1A")
    data_align = Alignment(vertical="top", wrap_text=True)
    data_brd   = Border(left=thin, right=thin, bottom=thin)
    yes_font   = Font(name="Consolas", size=10, color="1E8449", bold=True)
    no_font    = Font(name="Consolas", size=10, color="C0392B", bold=True)

    for r_idx, job in enumerate(jobs, 2):
        ws.append([job.to_row()[c] for c in COLUMNS])
        fill = PatternFill("solid", fgColor=row_fills[r_idx % len(row_fills)])
        for c_idx, cell in enumerate(ws[r_idx], 1):
            cell.fill = fill; cell.border = data_brd; cell.alignment = data_align
            cell.font = (yes_font if cell.value == "Yes" else no_font) if c_idx == 7 else data_font

    col_widths = [18, 50, 50, 12, 30, 40, 10, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    inst = wb.create_sheet("Instructions")
    inst.column_dimensions["A"].width = 80
    lines = [
        "BIM File Sync Automator — Job Export Format", "",
        "COLUMN GUIDE:",
        "  job_name     : Any label for the job",
        "  sources      : Full folder paths separated by  |",
        "  destination  : Single destination folder path",
        "  exec_time    : 24-hour time  HH:MM  (e.g.  09:00)",
        "  days         : Days separated by  |  — Mon Tue Wed Thu Fri Sat Sun",
        "  pick_list    : BIM codes separated by  |   Example:  WHZ | WHS",
        "  enabled      : Yes  or  No",
        "  filter_date  : YYYY-MM-DD or blank for all files",
        "  auto_today   : Yes = use today's date at run time",
        "",
        "TIPS:",
        "  • Pick-list: matches files containing BIM codes in their filename.",
        "  • Files copied: .nwc  .ifc  .xlsx",
        "  • Existing destination files are always overwritten.",
    ]
    for row_i, line in enumerate(lines, 1):
        cell = inst.cell(row=row_i, column=1, value=line)
        cell.font = Font(name="Arial", size=11,
                         bold=(line.startswith("BIM") or line.endswith(":")),
                         color="1A6FA8" if line.startswith("BIM") else "1A1A1A")
    wb.save(filepath)


def export_csv(jobs: List[SyncJob], filepath: str):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for job in jobs:
            w.writerow(job.to_row())


def import_excel(filepath: str) -> List[dict]:
    if not OPENPYXL_OK:
        raise ImportError("openpyxl not installed.")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Sync Jobs"] if "Sync Jobs" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(h).strip().lower().replace("\n","").replace(" ","_") if h else "" for h in rows[0]]
    return [{headers[i]: (str(row[i]).strip() if row[i] is not None else "")
             for i in range(min(len(headers), len(row)))}
            for row in rows[1:] if not all(v is None for v in row)]


def import_csv(filepath: str) -> List[dict]:
    with open(filepath, newline="", encoding="utf-8") as f:
        return [{k.strip().lower().replace(" ","_"): v.strip()
                 for k, v in row.items()} for row in csv.DictReader(f)]


# ── Sources Manager Dialog ─────────────────────────────────────────────────────
class SourcesDialog(ctk.CTkToplevel):
    def __init__(self, master, job: SyncJob):
        super().__init__(master)
        self.job = job; self.result = None
        self.title(f"Source Folders — {job.label}")
        self.geometry("640x480"); self.minsize(520, 360)
        self.configure(fg_color="#d6e8f7")
        self.grab_set(); self.lift(); self.focus_force()
        self._paths: List[str] = list(job.sources)
        self._rows: List[ctk.CTkFrame] = []
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(16, 6))
        ctk.CTkLabel(hdr, text=f"Source folders for {self.job.label}",
                     font=("Arial", 13, "bold"), text_color=COLORS["text"]).pack(side="left")
        ctk.CTkButton(hdr, text="＋  Add Source", width=130, height=30,
                      fg_color=COLORS["teal"], hover_color=COLORS["teal_h"],
                      command=self._add_source).pack(side="right")
        ctk.CTkLabel(self,
                     text="All sources scanned — .nwc  .ifc  .xlsx files copied to destination.",
                     font=("Arial", 11), text_color=COLORS["text_dim"]).pack(padx=20, anchor="w", pady=(0, 6))
        self.scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["src_bg"], corner_radius=8)
        self.scroll.pack(fill="both", expand=True, padx=20, pady=4)
        self.lbl_count = ctk.CTkLabel(self, text="", font=("Arial", 11), text_color=COLORS["text_dim"])
        self.lbl_count.pack(padx=20, anchor="w", pady=2)
        for path in self._paths:
            self._add_row(path)
        self._refresh_count()
        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(6, 16))
        ctk.CTkButton(btn_row, text="✔  Apply", height=36,
                      fg_color=COLORS["green"], hover_color=COLORS["green_h"],
                      command=self._apply).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ctk.CTkButton(btn_row, text="Cancel", height=36,
                      fg_color=COLORS["dim"], command=self._cancel).pack(side="left", expand=True, fill="x")

    def _add_source(self):
        path = filedialog.askdirectory(parent=self)
        if not path: return
        if path in self._paths:
            messagebox.showinfo("Duplicate", f"Already added:\n{path}", parent=self); return
        self._paths.append(path); self._add_row(path); self._refresh_count()

    def _add_row(self, path: str):
        idx = len(self._rows)
        row = ctk.CTkFrame(self.scroll, fg_color="#dce8f8", corner_radius=6)
        row.pack(fill="x", pady=3, padx=4)
        row.columnconfigure(1, weight=1)
        ctk.CTkLabel(row, text=f"#{idx+1}", width=28, font=("Consolas", 11, "bold"),
                     text_color=COLORS["teal"]).grid(row=0, column=0, padx=(8, 4), pady=8)
        ctk.CTkLabel(row, text=path, anchor="w", font=("Consolas", 10),
                     text_color=COLORS["text"]).grid(row=0, column=1, sticky="ew", padx=4)
        ctk.CTkButton(row, text="✕", width=28, height=26,
                      fg_color=COLORS["red"], hover_color=COLORS["red_h"],
                      font=("Arial", 11, "bold"),
                      command=lambda p=path, r=row: self._remove_row(p, r)).grid(
                          row=0, column=2, padx=(4, 8), pady=6)
        self._rows.append(row)

    def _remove_row(self, path: str, row: ctk.CTkFrame):
        if path in self._paths: self._paths.remove(path)
        row.destroy()
        self._rows = [r for r in self._rows if r.winfo_exists()]
        for i, r in enumerate(self._rows):
            w = r.winfo_children()
            if w: w[0].configure(text=f"#{i+1}")
        self._refresh_count()

    def _refresh_count(self):
        n = len(self._paths)
        self.lbl_count.configure(text=f"{n} source folder{'s' if n!=1 else ''} configured")

    def _apply(self):
        self.result = list(self._paths); self.grab_release(); self.destroy()

    def _cancel(self):
        self.grab_release(); self.destroy()


# ── Pick-list Editor Dialog ────────────────────────────────────────────────────
class PickListDialog(ctk.CTkToplevel):
    def __init__(self, master, job: SyncJob):
        super().__init__(master)
        self.job = job; self.result = None
        self.title(f"Pick-list — {job.label}")
        self.geometry("520x450"); self.resizable(False, False)
        self.configure(fg_color="#d6e8f7")
        self.grab_set(); self.lift(); self.focus_force()
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _build(self):
        ctk.CTkLabel(self, text=f"Pick-list codes for {self.job.label}",
                     font=("Arial", 13, "bold"),
                     text_color=COLORS["text"]).pack(pady=(16, 4), padx=20, anchor="w")
        ctk.CTkLabel(self, text="Standard codes — tick to include:",
                     font=("Arial", 11), text_color=COLORS["text_dim"]).pack(padx=20, anchor="w")

        cb_frame = ctk.CTkFrame(self, fg_color=COLORS["pick_bg"], corner_radius=8)
        cb_frame.pack(fill="x", padx=20, pady=6)
        current_set: Set[str] = {c.upper() for c in self.job.pick_list}
        self.cb_vars: dict[str, ctk.BooleanVar] = {}
        for i, code in enumerate(DEFAULT_PICKS):
            var = ctk.BooleanVar(value=(code in current_set))
            ctk.CTkCheckBox(cb_frame, text=code, variable=var,
                            font=("Consolas", 12, "bold"),
                            text_color=COLORS["text"],
                            checkbox_width=16, checkbox_height=16,
                            ).grid(row=i//4, column=i%4, padx=14, pady=8, sticky="w")
            self.cb_vars[code] = var

        qa = ctk.CTkFrame(self, fg_color="transparent")
        qa.pack(fill="x", padx=20, pady=(0, 4))
        ctk.CTkButton(qa, text="Select All", width=90, height=26,
                      command=lambda: self._set_all_codes(True)).pack(side="left", padx=(0, 6))
        ctk.CTkButton(qa, text="Clear All",  width=90, height=26, fg_color=COLORS["dim"],
                      command=lambda: self._set_all_codes(False)).pack(side="left")

        ctk.CTkLabel(self, text="Extra / custom codes (comma-separated):",
                     font=("Arial", 11), text_color=COLORS["text_dim"]).pack(padx=20, anchor="w", pady=(6, 2))
        default_set  = set(DEFAULT_PICKS)
        custom_codes = [c for c in self.job.pick_list if c.upper() not in default_set]
        self.custom_entry = ctk.CTkEntry(self, font=("Consolas", 11), height=32,
                                          fg_color="#daeaf7", text_color=COLORS["text"],
                                          border_color=COLORS["border"])
        self.custom_entry.insert(0, ", ".join(custom_codes))
        self.custom_entry.pack(fill="x", padx=20, pady=(0, 10))

        self.lbl_preview = ctk.CTkLabel(self, text="", font=("Arial", 10),
                                         text_color=COLORS["text_dim"], wraplength=460, anchor="w")
        self.lbl_preview.pack(padx=20, fill="x")
        self._update_preview()
        for var in self.cb_vars.values():
            var.trace_add("write", lambda *_: self._update_preview())
        self.custom_entry.bind("<KeyRelease>", lambda _: self._update_preview())

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(8, 16))
        ctk.CTkButton(btn_row, text="✔  Apply", height=36,
                      fg_color=COLORS["green"], hover_color=COLORS["green_h"],
                      command=self._apply).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ctk.CTkButton(btn_row, text="Cancel", height=36, fg_color=COLORS["dim"],
                      command=self._cancel).pack(side="left", expand=True, fill="x")

    def _set_all_codes(self, value: bool):
        for v in self.cb_vars.values():
            v.set(value)

    def _collect(self) -> List[str]:
        codes = [c for c, v in self.cb_vars.items() if v.get()]
        for c in self.custom_entry.get().split(","):
            c = c.strip().upper()
            if c and c not in codes: codes.append(c)
        return codes

    def _update_preview(self):
        codes = self._collect()
        self.lbl_preview.configure(
            text=("Active: " + ", ".join(codes)) if codes else "⚠ No codes — job will copy nothing.")

    def _apply(self):
        codes = self._collect()
        if not codes:
            messagebox.showwarning("Pick-list", "Select at least one code.", parent=self); return
        self.result = codes; self.grab_release(); self.destroy()

    def _cancel(self):
        self.grab_release(); self.destroy()


# ── Job Row Widget ─────────────────────────────────────────────────────────────
class JobRow(ctk.CTkFrame):
    def __init__(self, master, job: SyncJob, app, **kwargs):
        super().__init__(master, fg_color=COLORS["row_bg"], corner_radius=8,
                         border_width=1, border_color=COLORS["border"], **kwargs)
        self.job = job; self.app = app
        self._build()

    def _build(self):
        self.columnconfigure(2, weight=1)
        self.columnconfigure(3, weight=1)

        ctk.CTkLabel(self, text=self.job.label, font=("Arial", 11, "bold"),
                     fg_color=self.job.color, corner_radius=6,
                     width=62, text_color="white",
                     ).grid(row=0, column=0, rowspan=6, padx=(10, 6), pady=10, sticky="ns")

        name_frame = ctk.CTkFrame(self, fg_color="transparent")
        name_frame.grid(row=0, column=1, columnspan=2, padx=4, pady=(10, 2), sticky="ew")
        ctk.CTkLabel(name_frame, text="Name:", font=("Arial", 11),
                     text_color=COLORS["text"]).pack(side="left")
        self.name_var = ctk.StringVar(value=self.job.job_name)
        ctk.CTkEntry(name_frame, textvariable=self.name_var, width=260,
                     font=("Arial", 11), placeholder_text="Job name (optional)",
                     fg_color="#daeaf7", text_color=COLORS["text"],
                     border_color=COLORS["border"]).pack(side="left", padx=(6, 0))
        self.name_var.trace_add("write", lambda *_: setattr(self.job, "job_name", self.name_var.get()))

        ctk.CTkButton(self, text="Sources", width=72, height=26, font=("Arial", 11),
                      fg_color=COLORS["teal"], hover_color=COLORS["teal_h"],
                      command=self._open_sources).grid(row=1, column=1, padx=4, pady=2, sticky="w")
        self.lbl_src = ctk.CTkLabel(self, text=self._sources_summary(), anchor="w",
                                     font=("Consolas", 10),
                                     text_color=COLORS["text_dim"] if not self.job.sources else COLORS["teal"])
        self.lbl_src.grid(row=1, column=2, padx=4, sticky="ew")

        ctk.CTkButton(self, text="Dest", width=72, height=26, font=("Arial", 11),
                      fg_color="#6c7a89", hover_color="#5c6a79",
                      command=self._pick_dest).grid(row=2, column=1, padx=4, pady=2, sticky="w")
        self.lbl_dst = ctk.CTkLabel(self, text=self.job.dest or "No destination selected",
                                     anchor="w", font=("Consolas", 10),
                                     text_color=COLORS["text_dim"] if not self.job.dest else COLORS["text"])
        self.lbl_dst.grid(row=2, column=2, padx=4, sticky="ew")

        ctk.CTkButton(self, text="Pick-list", width=72, height=26, font=("Arial", 11),
                      fg_color=COLORS["purple"], hover_color="#5b2c6f",
                      command=self._open_picklist).grid(row=3, column=1, padx=4, pady=(2, 2), sticky="w")
        self.lbl_picks = ctk.CTkLabel(self, text=self._picks_summary(), anchor="w",
                                       font=("Consolas", 10), text_color=COLORS["purple"])
        self.lbl_picks.grid(row=3, column=2, padx=4, pady=(2, 10), sticky="ew")

        time_frame = ctk.CTkFrame(self, fg_color="transparent")
        time_frame.grid(row=1, column=3, padx=8, pady=2, sticky="ew")
        ctk.CTkLabel(time_frame, text="Time:", font=("Arial", 11),
                     text_color=COLORS["text"]).pack(side="left")
        self.time_var = ctk.StringVar(value=self.job.exec_time)
        ctk.CTkEntry(time_frame, textvariable=self.time_var, width=72,
                     font=("Consolas", 12), placeholder_text="HH:MM",
                     fg_color="#daeaf7", text_color=COLORS["text"],
                     border_color=COLORS["border"]).pack(side="left", padx=(4, 0))
        self.time_var.trace_add("write", lambda *_: setattr(self.job, "exec_time", self.time_var.get()))
        ctk.CTkLabel(time_frame, text="(24h  e.g. 09:00 / 14:30)",
                     font=("Arial", 9), text_color=COLORS["text_dim"]).pack(side="left", padx=(6, 0))

        days_frame = ctk.CTkFrame(self, fg_color="transparent")
        days_frame.grid(row=2, column=3, padx=8, pady=2, sticky="ew")
        self.day_vars: dict[str, ctk.BooleanVar] = {}
        for day in ALL_DAYS:
            var = ctk.BooleanVar(value=(day in self.job.days))
            ctk.CTkCheckBox(days_frame, text=day, variable=var, width=46,
                            font=("Arial", 10), text_color=COLORS["text"],
                            checkbox_width=14, checkbox_height=14,
                            command=self._on_days_change).pack(side="left", padx=2)
            self.day_vars[day] = var

        date_frame = ctk.CTkFrame(self, fg_color=COLORS["pick_bg"], corner_radius=6)
        date_frame.grid(row=3, column=3, rowspan=2, padx=8, pady=(2, 10), sticky="ew")

        ctk.CTkLabel(date_frame, text="Modified after:",
                     font=("Arial", 10, "bold"), text_color=COLORS["text"]).pack(side="left", padx=(8, 4))

        self.date_var = ctk.StringVar(value=self.job.filter_date)
        self.date_entry = ctk.CTkEntry(
            date_frame, textvariable=self.date_var, width=100,
            font=("Consolas", 11), placeholder_text="YYYY-MM-DD",
            fg_color="#daeaf7", text_color=COLORS["text"],
            border_color=COLORS["border"])
        self.date_entry.pack(side="left", padx=4)
        self.date_var.trace_add("write", self._on_date_change)

        ctk.CTkButton(date_frame, text="Today", width=52, height=24,
                      font=("Arial", 10), fg_color=COLORS["blue"], hover_color=COLORS["blue_h"],
                      command=self._set_today).pack(side="left", padx=2)
        ctk.CTkButton(date_frame, text="Yesterday", width=72, height=24,
                      font=("Arial", 10), fg_color=COLORS["teal"], hover_color=COLORS["teal_h"],
                      command=self._set_yesterday).pack(side="left", padx=2)
        ctk.CTkButton(date_frame, text="Clear", width=48, height=24,
                      font=("Arial", 10), fg_color=COLORS["dim"], hover_color="#4a6a8a",
                      command=self._clear_date).pack(side="left", padx=(2, 8))

        self.auto_today_var = ctk.BooleanVar(value=self.job.auto_today)
        self.chk_auto = ctk.CTkCheckBox(
            date_frame, text="Auto Today", variable=self.auto_today_var,
            font=("Arial", 10, "bold"), text_color=COLORS["green"],
            checkbox_width=14, checkbox_height=14,
            command=self._on_auto_today_change)
        self.chk_auto.pack(side="left", padx=(0, 8))

        self.lbl_date_status = ctk.CTkLabel(
            date_frame, text=self._date_status_text(),
            font=("Arial", 9), text_color=COLORS["orange"])
        self.lbl_date_status.pack(side="left", padx=4)

        self._apply_auto_today_state()

        stats_frame = ctk.CTkFrame(self, fg_color="transparent")
        stats_frame.grid(row=0, column=4, rowspan=6, padx=8, sticky="ns")
        self.lbl_ok   = ctk.CTkLabel(stats_frame, text="✔ 0", text_color=COLORS["green"], font=("Arial", 11))
        self.lbl_ok.pack(anchor="w")
        self.lbl_fail = ctk.CTkLabel(stats_frame, text="✘ 0", text_color=COLORS["red"],   font=("Arial", 11))
        self.lbl_fail.pack(anchor="w")
        self.lbl_last = ctk.CTkLabel(stats_frame, text="—",   text_color=COLORS["text_dim"], font=("Arial", 9))
        self.lbl_last.pack(anchor="w")

        ctrl_frame = ctk.CTkFrame(self, fg_color="transparent")
        ctrl_frame.grid(row=0, column=5, rowspan=6, padx=(4, 10), sticky="ns")
        self.enable_var = ctk.BooleanVar(value=self.job.enabled)
        ctk.CTkCheckBox(ctrl_frame, text="On", variable=self.enable_var, width=42,
                        font=("Arial", 11), text_color=COLORS["text"],
                        command=lambda: setattr(self.job, "enabled", self.enable_var.get()),
                        ).pack(pady=(12, 6))
        ctk.CTkButton(ctrl_frame, text="✕", width=32, height=26,
                      fg_color=COLORS["red"], hover_color=COLORS["red_h"],
                      font=("Arial", 12, "bold"),
                      command=lambda: self.app.remove_job(self.job)).pack()

    def _sources_summary(self) -> str:
        n = len(self.job.sources)
        if n == 0: return "No sources added"
        first = self.job.sources[0]
        trimmed = first if len(first) <= 45 else "…" + first[-44:]
        return trimmed if n == 1 else f"{trimmed}  (+{n-1} more)"

    def _picks_summary(self) -> str:
        picks = self.job.pick_list
        if not picks: return "⚠ No codes assigned"
        shown = ", ".join(picks[:6])
        return shown + (f"  (+{len(picks)-6} more)" if len(picks) > 6 else "")

    def _open_sources(self):
        dlg = SourcesDialog(self, self.job)
        self.wait_window(dlg)
        if dlg.result is not None:
            self.job.sources = dlg.result
            self.lbl_src.configure(text=self._sources_summary(),
                                    text_color=COLORS["teal"] if dlg.result else COLORS["text_dim"])
            self.app._log(f"[{self.job.label}] Sources updated: {len(dlg.result)} folder(s).")

    def _pick_dest(self):
        path = filedialog.askdirectory()
        if path:
            self.job.dest = path
            self.lbl_dst.configure(text=path, text_color=COLORS["text"])

    def _open_picklist(self):
        dlg = PickListDialog(self, self.job)
        self.wait_window(dlg)
        if dlg.result is not None:
            self.job.pick_list = dlg.result
            self.lbl_picks.configure(text=self._picks_summary())
            self.app._log(f"[{self.job.label}] Pick-list → {', '.join(dlg.result)}")

    def _on_days_change(self):
        self.job.days = [d for d, v in self.day_vars.items() if v.get()]

    def _on_date_change(self, *_):
        self.job.filter_date = self.date_var.get().strip()
        self.lbl_date_status.configure(text=self._date_status_text())

    def _on_auto_today_change(self):
        self.job.auto_today = self.auto_today_var.get()
        self._apply_auto_today_state()
        self.lbl_date_status.configure(text=self._date_status_text())

    def _apply_auto_today_state(self):
        if self.job.auto_today:
            self.date_entry.configure(state="disabled", fg_color="#c8dcea")
        else:
            self.date_entry.configure(state="normal", fg_color="#daeaf7")

    def refresh_date_display(self):
        if self.job.auto_today:
            self.date_var.set(self.job.filter_date)
        self.lbl_date_status.configure(text=self._date_status_text())

    def _set_today(self):
        self.auto_today_var.set(False)
        self.job.auto_today = False
        self._apply_auto_today_state()
        d = datetime.now().strftime("%Y-%m-%d")
        self.date_var.set(d)

    def _set_yesterday(self):
        from datetime import timedelta
        self.auto_today_var.set(False)
        self.job.auto_today = False
        self._apply_auto_today_state()
        d = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        self.date_var.set(d)

    def _clear_date(self):
        self.auto_today_var.set(False)
        self.job.auto_today = False
        self._apply_auto_today_state()
        self.date_var.set("")

    def _date_status_text(self) -> str:
        if self.job.auto_today:
            return f"🔄 Auto Today — will use today's date at run time"
        fd = self.job.filter_date.strip()
        if not fd:
            return "No date filter — all files"
        try:
            datetime.strptime(fd, "%Y-%m-%d")
            return f"✔ Only files modified on/after {fd}"
        except ValueError:
            return "⚠ Invalid date (use YYYY-MM-DD)"

    def refresh_stats(self):
        self.lbl_ok.configure(text=f"✔ {self.job.copied}")
        self.lbl_fail.configure(text=f"✘ {self.job.failed}")
        self.lbl_last.configure(text=self.job.last_run)


# ── Main Application ───────────────────────────────────────────────────────────
class BIMAutomatorApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("BIM File Sync Automator v8.5")
        self.geometry("1120x900")
        self.minsize(940, 720)
        self.configure(fg_color="#d6e8f7")

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.jobs:       List[SyncJob] = []
        self.job_rows:   List[JobRow]  = []
        self.is_running: bool          = False
        self._sched_lock = threading.Lock()
        self._next_id    = 0

        self._build_ui()

    def _build_ui(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=24, pady=(16, 4))
        ctk.CTkLabel(hdr, text="BIM File Sync Automator",
                     font=("Arial", 22, "bold"), text_color=COLORS["blue"]).pack(side="left")
        ctk.CTkLabel(hdr, text="v8.5", font=("Arial", 11),
                     text_color=COLORS["text_dim"]).pack(side="left", padx=(6, 0), pady=(6, 0))
        ctk.CTkLabel(hdr, text="Prepared by Ahmed Khalaf — BIM Manager",
                     font=("Arial", 11, "italic"),
                     text_color=COLORS["text_dim"]).pack(side="left", padx=(14, 0), pady=(6, 0))
        self.lbl_status = ctk.CTkLabel(hdr, text="● Idle", font=("Arial", 13),
                                        text_color=COLORS["text_dim"])
        self.lbl_status.pack(side="right")

        badge = ctk.CTkFrame(self, fg_color=COLORS["pick_bg"], corner_radius=6)
        badge.pack(fill="x", padx=24, pady=(0, 4))
        ctk.CTkLabel(badge,
                     text="Pick-list:  .nwc  .ifc  .xlsx  (matched by BIM code in filename) — existing files overwritten",
                     font=("Consolas", 11, "bold"), text_color=COLORS["blue"]).pack(
                         side="left", padx=14, pady=6)

        io_bar = ctk.CTkFrame(self, fg_color=COLORS["io_bar"], corner_radius=8,
                               border_width=1, border_color=COLORS["border"])
        io_bar.pack(fill="x", padx=24, pady=(0, 8))

        ctk.CTkLabel(io_bar, text="Import / Export:", font=("Arial", 12, "bold"),
                     text_color=COLORS["blue"]).pack(side="left", padx=(14, 10), pady=8)

        ctk.CTkButton(io_bar, text="⬇  Export Excel", width=130, height=32,
                      fg_color=COLORS["gold"], hover_color=COLORS["gold_h"],
                      font=("Arial", 11, "bold"),
                      command=self._export_excel).pack(side="left", padx=4, pady=6)
        ctk.CTkButton(io_bar, text="⬇  Export CSV", width=120, height=32,
                      fg_color="#4a7c59", hover_color="#3a6447",
                      font=("Arial", 11, "bold"),
                      command=self._export_csv).pack(side="left", padx=4, pady=6)

        ctk.CTkLabel(io_bar, text="│", text_color=COLORS["border"]).pack(side="left", padx=6)

        ctk.CTkButton(io_bar, text="⬆  Import & Merge", width=145, height=32,
                      fg_color=COLORS["blue"], hover_color=COLORS["blue_h"],
                      font=("Arial", 11, "bold"),
                      command=lambda: self._import(replace=False)).pack(side="left", padx=4, pady=6)
        ctk.CTkButton(io_bar, text="⬆  Import & Replace", width=155, height=32,
                      fg_color=COLORS["orange"], hover_color="#b84600",
                      font=("Arial", 11, "bold"),
                      command=lambda: self._import(replace=True)).pack(side="left", padx=4, pady=6)

        ctk.CTkLabel(io_bar, text="Supports .xlsx and .csv", font=("Arial", 10),
                     text_color=COLORS["text_dim"]).pack(side="right", padx=14)

        jlf = ctk.CTkFrame(self, fg_color="transparent")
        jlf.pack(fill="x", padx=24, pady=(4, 2))
        ctk.CTkLabel(jlf, text="Sync Jobs", font=("Arial", 14, "bold"),
                     text_color=COLORS["text"]).pack(side="left")
        ctk.CTkButton(jlf, text="＋  Add Job", width=110, height=30,
                      fg_color=COLORS["green"], hover_color=COLORS["green_h"],
                      command=self.add_job).pack(side="right")

        self.jobs_scroll = ctk.CTkScrollableFrame(self, height=420,
                                                   fg_color=COLORS["frame_bg"],
                                                   border_width=1, border_color=COLORS["border"])
        self.jobs_scroll.pack(fill="both", expand=True, padx=24, pady=4)

        stats_bar = ctk.CTkFrame(self, fg_color="transparent")
        stats_bar.pack(fill="x", padx=24, pady=4)
        self.lbl_g_ok   = ctk.CTkLabel(stats_bar, text="Total ✔ Copied: 0",
                                        text_color=COLORS["green"], font=("Arial", 12))
        self.lbl_g_ok.pack(side="left", padx=12)
        self.lbl_g_fail = ctk.CTkLabel(stats_bar, text="Total ✘ Failed: 0",
                                        text_color=COLORS["red"],   font=("Arial", 12))
        self.lbl_g_fail.pack(side="left", padx=12)
        self.lbl_g_last = ctk.CTkLabel(stats_bar, text="Last global run: —",
                                        text_color=COLORS["text_dim"], font=("Arial", 11))
        self.lbl_g_last.pack(side="right", padx=12)

        self.log_output = ctk.CTkTextbox(self, height=120, font=("Consolas", 11),
                                          fg_color="#daeaf7", text_color=COLORS["text"],
                                          border_width=1, border_color=COLORS["border"])
        self.log_output.pack(fill="x", padx=24, pady=6)
        self._log("System ready — add jobs or import from Excel / CSV.")
        self._log(f"Log file: {LOG_PATH}")

        btn_bar = ctk.CTkFrame(self, fg_color="transparent")
        btn_bar.pack(fill="x", padx=24, pady=(4, 16))

        self.btn_run = ctk.CTkButton(
            btn_bar, text="▶  Run All Jobs Now", height=46,
            font=("Arial", 15, "bold"),
            fg_color=COLORS["green"], hover_color=COLORS["green_h"],
            command=self.handle_execution)
        self.btn_run.pack(side="left", expand=True, fill="x", padx=(0, 6))

        self.btn_sched = ctk.CTkButton(
            btn_bar, text="⏱  Start Scheduler", height=46, width=160,
            font=("Arial", 13, "bold"),
            fg_color=COLORS["blue"], hover_color=COLORS["blue_h"],
            command=self.toggle_scheduler)
        self.btn_sched.pack(side="left", padx=6)

        ctk.CTkButton(btn_bar, text="Clear Log", height=46, width=100,
                      fg_color="#6c7a89", hover_color="#5c6a79",
                      command=self._clear_log).pack(side="left", padx=6)
        ctk.CTkButton(btn_bar, text="Open Log", height=46, width=100,
                      fg_color="#6c7a89", hover_color="#5c6a79",
                      command=self._open_log).pack(side="left", padx=6)
        ctk.CTkButton(btn_bar, text="⟲  Reset All", height=46, width=120,
                      font=("Arial", 12, "bold"),
                      fg_color=COLORS["red"], hover_color=COLORS["red_h"],
                      command=self._reset_all_settings).pack(side="left", padx=(6, 0))

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", padx=24, pady=(0, 10))
        ctk.CTkLabel(footer,
                     text="BIM File Sync Automator v8.5   •   Prepared by Ahmed Khalaf — BIM Manager",
                     font=("Arial", 10), text_color=COLORS["text_dim"]).pack(side="right")

        self.add_job()

    def add_job(self, job: Optional[SyncJob] = None) -> "JobRow":
        if job is None:
            job = SyncJob(job_id=self._next_id, pick_list=list(DEFAULT_PICKS))
        else:
            job.job_id = self._next_id
        self._next_id += 1
        self.jobs.append(job)
        row = JobRow(self.jobs_scroll, job, self)
        row.pack(fill="x", pady=4, padx=4)
        self.job_rows.append(row)
        return row

    def remove_job(self, job: SyncJob):
        if len(self.jobs) == 1:
            messagebox.showinfo("Remove Job", "At least one job must remain."); return
        idx = next((i for i, j in enumerate(self.jobs) if j is job), None)
        if idx is None: return
        self.jobs.pop(idx); self.job_rows.pop(idx).destroy()
        self._log(f"Removed {job.label}.")

    def _clear_all_jobs(self):
        for row in list(self.job_rows): row.destroy()
        self.jobs.clear(); self.job_rows.clear()

    def _reset_all_settings(self):
        """Wipe all jobs, stats, and the log, then start fresh with one default job."""
        if not messagebox.askyesno(
            "Reset All Settings",
            f"This will remove all {len(self.jobs)} job(s), reset all counters, "
            "and clear the on-screen log.\n\n"
            "Your saved log file and any exported Excel/CSV files are NOT affected.\n\n"
            "Continue?",
        ):
            return

        # Stop the scheduler if it's running
        if self.is_running:
            self.is_running = False
            with self._sched_lock:
                schedule.clear()
            self.btn_sched.configure(text="⏱  Start Scheduler",
                                     fg_color=COLORS["blue"], hover_color=COLORS["blue_h"])

        # Wipe jobs and reset id counter
        self._clear_all_jobs()
        self._next_id = 0

        # Reset global stats
        self.lbl_g_ok.configure(text="Total ✔ Copied: 0")
        self.lbl_g_fail.configure(text="Total ✘ Failed: 0")
        self.lbl_g_last.configure(text="Last global run: —")
        self._set_status("● Idle", COLORS["text_dim"])

        # Clear the on-screen log and start fresh
        self.log_output.delete("0.0", "end")
        self._log("All settings reset — starting fresh with one default job.")

        # Add a single clean default job
        self.add_job()

    def _export_excel(self):
        if not self.jobs:
            messagebox.showwarning("Export", "No jobs to export."); return
        if not OPENPYXL_OK:
            messagebox.showerror("Missing Library",
                                  "openpyxl is required.\n\nRun:  pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel Workbook", "*.xlsx")],
                                             initialfile="bim_sync_jobs.xlsx")
        if not path: return
        try:
            export_excel(self.jobs, path)
            self._log(f"Exported {len(self.jobs)} job(s) → {path}")
            messagebox.showinfo("Export Complete", f"Exported {len(self.jobs)} job(s) to:\n{path}")
        except Exception as exc:
            self._log(f"Export error: {exc}"); messagebox.showerror("Export Error", str(exc))

    def _export_csv(self):
        if not self.jobs:
            messagebox.showwarning("Export", "No jobs to export."); return
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                             filetypes=[("CSV file", "*.csv")],
                                             initialfile="bim_sync_jobs.csv")
        if not path: return
        try:
            export_csv(self.jobs, path)
            self._log(f"Exported {len(self.jobs)} job(s) → {path}")
            messagebox.showinfo("Export Complete", f"Exported {len(self.jobs)} job(s) to:\n{path}")
        except Exception as exc:
            self._log(f"Export error: {exc}"); messagebox.showerror("Export Error", str(exc))

    def _import(self, replace: bool):
        path = filedialog.askopenfilename(
            filetypes=[("Excel / CSV", "*.xlsx *.csv"),
                       ("Excel Workbook", "*.xlsx"), ("CSV file", "*.csv")])
        if not path: return
        try:
            ext  = os.path.splitext(path)[1].lower()
            rows = import_excel(path) if ext == ".xlsx" else import_csv(path) if ext == ".csv" else None
            if rows is None:
                messagebox.showerror("Import", "Unsupported file type."); return
            if not rows:
                messagebox.showwarning("Import", "No data rows found."); return
            new_jobs = [SyncJob.from_row(r, 0) for r in rows]
            if replace:
                if not messagebox.askyesno("Import & Replace",
                    f"Remove all {len(self.jobs)} current job(s) and load "
                    f"{len(new_jobs)} from file?\n\nContinue?"): return
                self._clear_all_jobs()
                self._log("All existing jobs cleared for replace-import.")
            for job in new_jobs:
                self.add_job(job)
                self._log(f"Imported: {job.label} | {len(job.sources)} src(s) | "
                          f"{', '.join(job.days)} {job.exec_time}")
            mode = "replaced with" if replace else "merged +"
            messagebox.showinfo("Import Complete",
                                f"Successfully {mode} {len(new_jobs)} job(s) from:\n{path}")
        except Exception as exc:
            self._log(f"Import error: {exc}"); messagebox.showerror("Import Error", str(exc))

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_output.insert("end", f"[{ts}] {msg}\n")
        self.log_output.see("end")
        logger.info(msg)

    def _clear_log(self):
        self.log_output.delete("0.0", "end"); self._log("Log cleared.")

    @staticmethod
    def _open_log():
        # Open the log from its writable location (LOCALAPPDATA), not the installation dir.
        path = LOG_PATH
        if os.path.exists(path):
            try:
                os.startfile(path)                       # Windows
            except AttributeError:
                import subprocess, sys
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.Popen([opener, path])
        else:
            messagebox.showinfo("Log", f"No log file found yet.\n\nExpected at:\n{path}")

    def _set_status(self, text: str, color: str):
        self.lbl_status.configure(text=text, text_color=color)

    def _run_job(self, job: SyncJob):
        if not job.enabled:
            self._log(f"[{job.label}] Skipped (disabled)."); return
        if not job.sources:
            self._log(f"[{job.label}] ERROR: No source folders."); return
        if not job.dest:
            self._log(f"[{job.label}] ERROR: No destination selected."); return
        if not job.pick_list:
            self._log(f"[{job.label}] ERROR: No pick-list codes selected."); return

        dest_real = os.path.abspath(job.dest)
        dest_long = _long_path(dest_real)          # long-path safe destination root
        # Diagnose the destination before trying to create it, so a disconnected
        # drive gives a clear reason instead of a cryptic OS error.
        ok_dest, why_dest = _diagnose_path(job.dest)
        if not ok_dest and "folder not found" not in why_dest:
            # Drive/share problem — creating folders won't help.
            self._log(f"[{job.label}] ERROR: Destination unavailable — {why_dest}  ({job.dest})")
            return
        try:
            os.makedirs(dest_long, exist_ok=True)
        except Exception as exc:
            self._log(f"[{job.label}] ERROR: Cannot create destination '{job.dest}' — {exc}"); return
        if not os.path.isdir(dest_long):
            self._log(f"[{job.label}] ERROR: Destination not accessible '{job.dest}'."); return

        upper_picks = [c.upper() for c in job.pick_list]

        if job.auto_today:
            job.filter_date = datetime.now().strftime("%Y-%m-%d")
            row = next((r for r in self.job_rows if r.job is job), None)
            if row:
                self.after(0, row.refresh_date_display)

        filter_dt = job.filter_datetime

        self._log(
            f"[{job.label}] ── Scan started | {len(job.sources)} src(s) | "
            f"Pick-list codes: {len(upper_picks)} | "
            + (f"modified on/after: {job.filter_date}" if filter_dt else "no date filter")
        )
        ok = fail = 0
        skipped_sources = 0

        for src_root in job.sources:
            ok_src, why_src = _diagnose_path(src_root)
            if not ok_src:
                skipped_sources += 1
                self._log(f"[{job.label}]   ⚠ Skipping source ({why_src}): {src_root}")
                continue
            src_root_real = _long_path(src_root)
            self._log(f"[{job.label}]   Scanning: {src_root}")

            for root, _, files in os.walk(src_root_real):
                for filename in files:
                    ext_lower = filename.lower()
                    src = os.path.join(root, filename)

                    matched = (
                        ext_lower.endswith(EXTENSIONS)
                        and upper_picks
                        and any(code in filename.upper() for code in upper_picks)
                    )
                    if not matched:
                        continue

                    if filter_dt is not None:
                        try:
                            file_mtime = datetime.fromtimestamp(os.path.getmtime(src))
                        except Exception:
                            file_mtime = datetime.min
                        if file_mtime < filter_dt:
                            self._log(
                                f"[{job.label}]   Skipped (modified "
                                f"{file_mtime.strftime('%Y-%m-%d')} < {job.filter_date}): {filename}"
                            )
                            continue

                    # Long-path safe on BOTH ends. shutil.copy2 overwrites any
                    # existing file at the destination by default.
                    dst = os.path.join(dest_long, filename)

                    try:
                        overwrote = os.path.exists(dst)
                        shutil.copy2(src, dst)
                        action = "Overwrote" if overwrote else "Copied"
                        self._log(f"[{job.label}]   ✔ {action}: {filename}")
                        ok += 1
                    except Exception as exc:
                        self._log(f"[{job.label}]   ✘ Failed: {filename} — {exc}")
                        fail += 1

        job.copied += ok
        job.failed += fail
        job.last_run = datetime.now().strftime("%H:%M:%S")
        if skipped_sources:
            self._log(
                f"[{job.label}] ⚠ {skipped_sources} of {len(job.sources)} source(s) "
                f"were unreachable and skipped — check the warnings above."
            )
        self._log(f"[{job.label}] ── Done. Copied {ok}, Failed {fail}")
        self.after(0, self._refresh_all_stats)

    def _refresh_all_stats(self):
        total_ok = total_fail = 0
        for row in self.job_rows:
            row.refresh_stats()
            total_ok += row.job.copied; total_fail += row.job.failed
        self.lbl_g_ok.configure(text=f"Total ✔ Copied: {total_ok}")
        self.lbl_g_fail.configure(text=f"Total ✘ Failed: {total_fail}")
        self.lbl_g_last.configure(text=f"Last global run: {datetime.now().strftime('%H:%M:%S')}")

    def _run_all_jobs(self):
        self._set_status("● Running", COLORS["orange"])
        for job in list(self.jobs): self._run_job(job)
        self._set_status("● Scheduled" if self.is_running else "● Idle",
                         COLORS["blue"] if self.is_running else COLORS["text_dim"])

    def handle_execution(self):
        if not self.jobs:
            messagebox.showwarning("No Jobs", "Add at least one sync job first."); return
        threading.Thread(target=self._run_all_jobs, daemon=True).start()

    def toggle_scheduler(self):
        if not self.is_running:
            errors = []
            for job in self.jobs:
                if not job.enabled: continue
                if not job.sources:
                    errors.append(f"{job.label}: no source folders.")
                if not job.dest:
                    errors.append(f"{job.label}: no destination.")
                if not validate_time(job.exec_time):
                    errors.append(f"{job.label}: invalid time '{job.exec_time}'.")
                if not job.days:
                    errors.append(f"{job.label}: no days selected.")
                if not job.pick_list:
                    errors.append(f"{job.label}: pick-list is empty.")
            if errors:
                messagebox.showerror("Scheduler Errors", "\n".join(errors)); return
            self.is_running = True
            self.btn_sched.configure(text="■  Stop Scheduler",
                                      fg_color=COLORS["red"], hover_color=COLORS["red_h"])
            self._set_status("● Scheduled", COLORS["blue"])
            threading.Thread(target=self._scheduler_loop, daemon=True).start()
            self._log("Scheduler started.")
        else:
            self.is_running = False
            with self._sched_lock: schedule.clear()
            self.btn_sched.configure(text="⏱  Start Scheduler",
                                      fg_color=COLORS["blue"], hover_color=COLORS["blue_h"])
            self._set_status("● Idle", COLORS["text_dim"])
            self._log("Scheduler stopped.")

    def _scheduler_loop(self):
        with self._sched_lock:
            schedule.clear()
            for job in self.jobs:
                if not job.enabled: continue
                for day in job.days:
                    day_attr = DAY_MAP.get(day)
                    if not day_attr: continue
                    try:
                        getattr(schedule.every(), day_attr).at(job.exec_time).do(
                            lambda j=job: threading.Thread(
                                target=self._run_job, args=(j,), daemon=True).start())
                        self._log(f"Scheduled {job.label} → {day} {job.exec_time}")
                    except Exception as exc:
                        self._log(f"Schedule error for {job.label}: {exc}")
        while self.is_running:
            with self._sched_lock: schedule.run_pending()
            time.sleep(10)


# ── Entry Point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = BIMAutomatorApp()
    app.mainloop()